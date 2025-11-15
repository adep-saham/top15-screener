import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile
import os

# ===================== PAGE CONFIG =====================
st.set_page_config(
    page_title="Top15 Smart Entry + 7D Protection",
    layout="wide"
)

# ===================== GOLD-BLACK THEME CSS =====================
st.markdown("""
    <style>
    /* Background */
    .main {
        background-color: #050608;
        color: #F8F3E7;
    }
    body {
        background-color: #050608;
    }
    /* Title & Subtitle */
    .title-text {
        font-size: 40px;
        font-weight: 900;
        color: #F5D26B;
        padding-bottom: 0px;
    }
    .subtitle-text {
        font-size: 17px;
        color: #D8D2C2;
        margin-top: -6px;
        font-weight: 400;
    }
    /* Section Title */
    .section-title {
        font-size: 22px;
        font-weight: 700;
        color: #F5D26B;
        padding-top: 18px;
        padding-bottom: 5px;
    }
    /* Upload Box */
    .upload-box {
        padding: 18px;
        background: linear-gradient(135deg, #1A1205 0%, #111317 60%, #1F1A0D 100%);
        border-radius: 12px;
        border: 1px solid #F5D26B33;
        color: #F3EBDD;
        font-size: 14px;
    }
    .upload-box b {
        color: #FAD97F;
    }
    /* Info Box */
    .info-box {
        padding: 10px 16px;
        background-color: #111317;
        border-radius: 8px;
        border-left: 3px solid #F5D26B;
        color: #E7E0CF;
        font-size: 13px;
        margin-top: 10px;
    }
    /* Footer */
    .footer-text {
        text-align: center;
        padding-top: 25px;
        padding-bottom: 10px;
        color: #A99C7A;
        font-size: 12px;
    }
    /* Dataframe tweak */
    .stDataFrame {
        background-color: #050608;
    }
    </style>
""", unsafe_allow_html=True)

# ===================== HEADER =====================
st.markdown("""
<div style='display:flex; align-items:center; gap:16px; margin-bottom:10px;'>
    <img src="https://img.icons8.com/fluency/96/combo-chart--v1.png" width="70">
    <div>
        <div class="title-text">Top 15 Smart Entry + 7D Momentum Protection</div>
        <div class="subtitle-text">
            Screener otomatis berbasis flow, bandarmology, dan momentum pendek dengan proteksi drawdown 1 minggu.
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ===================== UPLOAD SECTION =====================
st.markdown("<div class='section-title'>📂 Upload File Screener</div>", unsafe_allow_html=True)

st.markdown("""
<div class='upload-box'>
<b>Upload 6 file utama dari Stockbit (CSV/XLSX):</b><br>
1️⃣ <b>1 Week Net Foreign Flow</b><br>
2️⃣ <b>1 Month Net Foreign Flow</b><br>
3️⃣ <b>Bandar Accumulation / Bandar Accummulation Uptrend</b><br>
4️⃣ <b>Frequency</b><br>
5️⃣ <b>High Volume Breakout</b><br>
6️⃣ <b>Reversal Signal</b><br><br>

<b>Opsional:</b> tambahkan file <b>7D Momentum Protection</b> (screener teknikal pendek yang sudah kamu buat di Stockbit).<br>
Saham yang lolos screener 7D akan mendapat prioritas di hasil Top 15.
</div>
""", unsafe_allow_html=True)

uploaded_files = st.file_uploader(
    "Upload file-file Stockbit di sini",
    type=["csv", "xlsx", "xls"],
    accept_multiple_files=True,
)

st.markdown("""
<div class="info-box">
💡 <b>Tips:</b> Pastikan kolom pertama di setiap file adalah <b>kode saham</b> (Symbol) dan ada kolom <b>Price</b>.
Nama file tidak harus persis, asalkan mengandung kata kunci: "1 Week", "1 Month", "Bandar", "Frequency", "High Volume", "Reversal", atau "7D".
</div>
""", unsafe_allow_html=True)

# ===================== UTILS =====================
def to_num(x):
    if pd.isna(x):
        return np.nan
    s = str(x).replace("Rp", "").replace("IDR", "").replace(",", "").replace(".JK", "")
    try:
        return float(s)
    except:
        return np.nan

def read_any(file):
    name = file.name.lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(file)
        except:
            file.seek(0)
            return pd.read_csv(file, sep=";")
    else:
        return pd.read_excel(file)

def norm_cols(df):
    df.columns = [c.lower().strip() for c in df.columns]
    return df

def canonical_label(name: str):
    n = name.lower()
    if "1 week" in n or "1w" in n: return "ff1w"
    if "1 month" in n or "1m" in n: return "ff1m"
    if "bandar" in n: return "bandar"
    if "frequency" in n or "freq" in n: return "frequency"
    if "high" in n and "volume" in n: return "hvb"
    if "reversal" in n or "revesal" in n: return "reversal"
    if "7d" in n or "momentum protection" in n: return "prot7d"
    return "other"

def color_rr(path):
    wb = load_workbook(path)
    ws = wb.active
    rr_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "RR":
            rr_col = i
            break
    if rr_col is None:
        wb.save(path)
        return

    for row in ws.iter_rows(min_row=2, min_col=rr_col, max_col=rr_col):
        for c in row:
            if c.value is None:
                continue
            try:
                val = float(c.value)
            except:
                continue
            if val >= 2:
                c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # hijau
            elif val >= 1.5:
                c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # kuning
            else:
                c.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # merah
    wb.save(path)

# ===================== MAIN LOGIC =====================

if uploaded_files and st.button("🚀 Proses Screener", use_container_width=True):
    st.markdown("<div class='section-title'>🔎 Langkah 1 — Baca & Gabungkan Data</div>", unsafe_allow_html=True)

    dfs = []
    signals = {k: set() for k in ["ff1w", "ff1m", "bandar", "frequency", "hvb", "reversal"]}
    prot7d_tickers = set()

    for f in uploaded_files:
        df = read_any(f)
        df = norm_cols(df)
        label = canonical_label(f.name)

        if "price" not in df.columns:
            st.warning(f"File **{f.name}** di-skip (tidak ada kolom 'Price').")
            continue

        first_col = df.columns[0]
        df["ticker"] = df[first_col].astype(str).str.upper().str.replace(".JK", "", regex=False)
        df["price"] = df["price"].map(to_num)

        dfs.append(df[["ticker", "price"]])

        if label in signals:
            signals[label] |= set(df["ticker"])
        elif label == "prot7d":
            prot7d_tickers |= set(df["ticker"])
            st.info(f"✅ 7D Momentum Protection terdeteksi dari file **{f.name}** ({len(prot7d_tickers)} ticker).")
        else:
            st.write(f"ℹ️ File **{f.name}** hanya dipakai untuk harga (bukan sinyal utama).")

    if not dfs:
        st.error("Tidak ada file valid yang punya kolom 'Price'. Cek kembali input.")
        st.stop()

    raw = pd.concat(dfs, ignore_index=True)
    agg = raw.groupby("ticker", as_index=False).agg({"price": "median"})

    sig = pd.DataFrame({"ticker": agg["ticker"]})
    for k in signals:
        sig[k] = sig["ticker"].isin(signals[k]).astype(int)

    sig["signal_count"] = sig[list(signals.keys())].sum(axis=1)

    def infer_cat(r):
        if r["frequency"] and r["hvb"]:
            return "Scalping"
        if r["frequency"] or r["hvb"]:
            return "Intraday"
        if r["ff1w"] or r["ff1m"] or r["bandar"] or r["reversal"]:
            return "Swing"
        return "Watchlist"

    sig["category"] = sig.apply(infer_cat, axis=1)

    base = sig.merge(agg, on="ticker", how="left")
    base = base[base["category"] != "Watchlist"].copy()

    st.write(f"Jumlah kandidat setelah kategori (Scalping/Intraday/Swing): **{len(base)}**")
    st.dataframe(base.head())

    if len(base) == 0:
        st.error("Tidak ada saham yang lolos kategori (Scalping/Intraday/Swing). Cek file sinyal.")
        st.stop()

    # ---------------- SMART ENTRY RANGE ----------------
    st.markdown("<div class='section-title'>📐 Langkah 2 — Hitung Entry, Stop, Target, & RR</div>", unsafe_allow_html=True)

    tick = 1
    risk = {"Scalping": 0.01, "Intraday": 0.02, "Swing": 0.035}
    rr_mult = {"Scalping": 2.5, "Intraday": 2.5, "Swing": 3.0}

    def ceil(x, t): return np.ceil(x / t) * t
    def floor(x, t): return np.floor(x / t) * t

    def smart_range(r):
        if pd.isna(r["price"]) or r["price"] == 0:
            return pd.Series({
                "entry_low": np.nan, "entry_high": np.nan, "entry_mid": np.nan,
                "stop": np.nan, "target": np.nan, "RR": np.nan,
                "entry_type": "N/A", "ladder": "N/A"
            })
        cat = r["category"]
        px = r["price"]
        rng = max(px * 0.02, 5)

        if cat == "Scalping":
            low, high = ceil(px * 1.01, tick), ceil(px * 1.015, tick)
            et = "Breakout Range"
            stop = floor(low * (1 - risk[cat]), tick)
        elif cat == "Intraday":
            low, high = ceil(px * 0.99, tick), ceil(px * 1.01, tick)
            et = "Retest Range"
            stop = floor(low - 0.3 * rng, tick)
        else:
            low, high = ceil(px * 0.98, tick), ceil(px * 0.995, tick)
            et = "MA20 Pullback"
            stop = floor(low * (1 - risk[cat]), tick)

        mid = (low + high) / 2
        target = ceil(mid + (mid - stop) * rr_mult[cat], tick)
        rr = round((target - mid) / (mid - stop), 2)
        ladder = (
            f"40%@{int(low)} | 20%@{int(low + (high - low) * 0.33)} | "
            f"25%@{int(low + (high - low) * 0.66)} | 15%@{int(high)}"
        )

        return pd.Series({
            "entry_low": low, "entry_high": high, "entry_mid": mid,
            "stop": stop, "target": target, "RR": rr,
            "entry_type": et, "ladder": ladder
        })

    plan = pd.concat([base, base.apply(smart_range, axis=1)], axis=1)

    # Flag proteksi 7 hari
    plan["prot7d"] = plan["ticker"].isin(prot7d_tickers).astype(int)
    st.write(f"Ticker dengan proteksi 7D: **{plan['prot7d'].sum()}**")

    # Filter ringan
    plan = plan[(plan["RR"] >= 1.8) & (plan["signal_count"] >= 2)].copy()

    # ---------------- SCORING ----------------
    st.markdown("<div class='section-title'>🏆 Langkah 3 — Ranking & Top 15</div>", unsafe_allow_html=True)

    plan["score_raw"] = plan["signal_count"] + plan["RR"]
    plan["score"] = plan["score_raw"] + 0.7 * plan["prot7d"]

    plan_sorted = plan.sort_values(["prot7d", "score", "RR"], ascending=[False, False, False])

    protected = plan_sorted[plan_sorted["prot7d"] == 1].head(15)
    if len(protected) < 15:
        remaining = plan_sorted[plan_sorted["prot7d"] == 0].head(15 - len(protected))
        top15 = pd.concat([protected, remaining], ignore_index=True)
    else:
        top15 = protected.copy()

    st.subheader("🏅 Top 15 Terpilih")
    st.dataframe(top15[[
        "ticker", "category", "entry_type",
        "entry_low", "entry_high", "target", "stop",
        "RR", "signal_count", "prot7d", "score"
    ]])

    # ---------------- SAVE EXCEL & DOWNLOAD ----------------
    st.markdown("<div class='section-title'>⬇️ Langkah 4 — Download Excel</div>", unsafe_allow_html=True)

    ts = datetime.now().strftime("%Y-%m-%d_%H%M")

    # Top15
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_top:
        top_path = tmp_top.name
    with pd.ExcelWriter(top_path, engine="openpyxl") as w:
        top15.to_excel(w, sheet_name="Top15", index=False)
    color_rr(top_path)
    with open(top_path, "rb") as f:
        top_bytes = f.read()
    os.remove(top_path)

    # All plan
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_plan:
        plan_path = tmp_plan.name
    with pd.ExcelWriter(plan_path, engine="openpyxl") as w:
        plan.to_excel(w, sheet_name="All", index=False)
    color_rr(plan_path)
    with open(plan_path, "rb") as f:
        plan_bytes = f.read()
    os.remove(plan_path)

    col1, col2 = st.columns(2)
    with col1:
        st.download_button(
            label="📥 Download Top 15 (Excel)",
            data=top_bytes,
            file_name=f"Top15_SmartColor_7D_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col2:
        st.download_button(
            label="📥 Download All Plan (Excel)",
            data=plan_bytes,
            file_name=f"All_SmartColor_7D_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ===================== FOOTER =====================
st.markdown("""
<div class='footer-text'>
Top15 Screener · Gold & Black Edition · Dibuat untuk membantu swing trader mengurangi drawdown mingguan.<br>
Powered by Streamlit.
</div>
""", unsafe_allow_html=True)
